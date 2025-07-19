import pandas as pd
import openpyxl
from openpyxl import load_workbook
import logging

class ExcelProcessor:
    """Service to process Excel files containing analyst estimates and financial data"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def process_excel(self, file_path):
        """
        Process Excel file to extract financial data and analyst estimates
        Returns structured data with line items, previous values, and estimates
        """
        try:
            # Load workbook with openpyxl to handle formatting
            workbook = load_workbook(file_path, data_only=True)
            
            # Try to find the most relevant worksheet
            worksheet = self._find_financial_worksheet(workbook)
            
            # Convert to pandas DataFrame for easier processing
            df = pd.DataFrame(worksheet.values)
            
            # Extract financial data
            financial_data = self._extract_financial_data(df)
            
            self.logger.info(f"Successfully processed Excel file: {file_path}")
            return financial_data
            
        except Exception as e:
            self.logger.error(f"Error processing Excel file {file_path}: {str(e)}")
            raise Exception(f"Failed to process Excel file: {str(e)}")
    
    def _find_financial_worksheet(self, workbook):
        """Find the most relevant worksheet containing financial data"""
        # Common worksheet names for financial data
        target_names = ['financials', 'estimates', 'data', 'summary', 'model']
        
        # First, try to find by name
        for sheet_name in workbook.sheetnames:
            if any(target in sheet_name.lower() for target in target_names):
                return workbook[sheet_name]
        
        # If no specific name found, use the first sheet
        return workbook.active
    
    def _extract_financial_data(self, df):
        """Extract and structure financial data from DataFrame"""
        financial_data = {
            'line_items': [],
            'metadata': {
                'total_rows': len(df),
                'total_columns': len(df.columns) if not df.empty else 0
            }
        }
        
        # Common financial line items to look for
        key_metrics = [
            'revenue', 'net revenue', 'total revenue', 'sales',
            'gross profit', 'gross margin',
            'ebitda', 'adjusted ebitda', 'operating income',
            'net income', 'net earnings',
            'eps', 'earnings per share', 'diluted eps',
            'free cash flow', 'operating cash flow',
            'total assets', 'total debt', 'shareholders equity'
        ]
        
        try:
            # Find header row (usually contains 'metric', 'line item', or period indicators)
            header_row_idx = self._find_header_row(df)
            
            if header_row_idx is not None:
                # Set header row as column names
                df.columns = df.iloc[header_row_idx]
                df = df.drop(df.index[header_row_idx]).reset_index(drop=True)
            
            # Process each row to find financial metrics
            for idx, row in df.iterrows():
                if pd.isna(row.iloc[0]) or str(row.iloc[0]).strip() == '':
                    continue
                
                line_item_name = str(row.iloc[0]).strip()
                
                # Check if this row contains a key financial metric
                if self._is_financial_metric(line_item_name, key_metrics):
                    line_item_data = self._extract_line_item_data(row, line_item_name)
                    if line_item_data:
                        financial_data['line_items'].append(line_item_data)
            
            # If no specific metrics found, extract first few numeric rows
            if not financial_data['line_items']:
                financial_data['line_items'] = self._extract_generic_data(df)
            
        except Exception as e:
            self.logger.warning(f"Error in detailed extraction, using generic method: {str(e)}")
            financial_data['line_items'] = self._extract_generic_data(df)
        
        return financial_data
    
    def _find_header_row(self, df):
        """Find the row that likely contains column headers"""
        for idx, row in df.iterrows():
            # Look for rows containing period indicators or common headers
            row_str = ' '.join([str(cell).lower() for cell in row if pd.notna(cell)])
            if any(indicator in row_str for indicator in ['q1', 'q2', 'q3', 'q4', '2024', '2025', 'previous', 'estimate', 'actual']):
                return idx
        return None
    
    def _is_financial_metric(self, line_item_name, key_metrics):
        """Check if a line item name matches known financial metrics"""
        line_item_lower = line_item_name.lower()
        return any(metric in line_item_lower for metric in key_metrics)
    
    def _extract_line_item_data(self, row, line_item_name):
        """Extract numeric data for a specific line item"""
        try:
            # Get all numeric values from the row (excluding the first column which is the name)
            numeric_values = []
            for value in row.iloc[1:]:
                if pd.notna(value) and str(value).strip() != '':
                    try:
                        # Try to convert to float, handling common formats
                        cleaned_value = str(value).replace(',', '').replace('$', '').replace('(', '-').replace(')', '')
                        numeric_val = float(cleaned_value)
                        numeric_values.append(numeric_val)
                    except (ValueError, TypeError):
                        continue
            
            if len(numeric_values) >= 2:
                return {
                    'line_item': line_item_name,
                    'previous_value': numeric_values[0] if len(numeric_values) > 0 else None,
                    'estimate': numeric_values[1] if len(numeric_values) > 1 else None,
                    'actual': None,  # Placeholder for actual values
                    'all_values': numeric_values
                }
        except Exception as e:
            self.logger.warning(f"Error extracting data for {line_item_name}: {str(e)}")
        
        return None
    
    def _extract_generic_data(self, df):
        """Extract data using a generic approach when specific metrics aren't found"""
        line_items = []
        
        try:
            # Take first 20 rows that have both text and numeric data
            count = 0
            for idx, row in df.iterrows():
                if count >= 20:
                    break
                
                if pd.isna(row.iloc[0]) or str(row.iloc[0]).strip() == '':
                    continue
                
                line_item_name = str(row.iloc[0]).strip()
                numeric_values = []
                
                # Extract numeric values
                for value in row.iloc[1:]:
                    if pd.notna(value):
                        try:
                            cleaned_value = str(value).replace(',', '').replace('$', '').replace('(', '-').replace(')', '')
                            numeric_val = float(cleaned_value)
                            numeric_values.append(numeric_val)
                        except (ValueError, TypeError):
                            continue
                
                if len(numeric_values) >= 1:
                    line_items.append({
                        'line_item': line_item_name,
                        'previous_value': numeric_values[0] if len(numeric_values) > 0 else None,
                        'estimate': numeric_values[1] if len(numeric_values) > 1 else numeric_values[0],
                        'actual': None,
                        'all_values': numeric_values
                    })
                    count += 1
        
        except Exception as e:
            self.logger.error(f"Error in generic data extraction: {str(e)}")
        
        return line_items
